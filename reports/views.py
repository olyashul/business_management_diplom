from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.utils import timezone
from django.db import models
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .models import Report
from sales.models import Sale
from products.models import Product
from staff.models import Employee, WorkShift


@login_required
def reports_dashboard(request):
    """Главная страница отчетов"""
    today = timezone.now().date()
    
    # Статистика для карточек
    today_sales = Sale.objects.filter(created_at__date=today, is_return=False)
    weekly_sales = Sale.objects.filter(
        created_at__date__gte=today - timedelta(days=7),
        is_return=False
    )
    
    from products.models import Category
    from staff.models import Employee
    
    context = {
        'today': today,
        'total_sales_today': today_sales.count(),
        'revenue_today': sum(sale.final_amount for sale in today_sales) if today_sales else 0,
        'total_sales_week': weekly_sales.count(),
        'revenue_week': sum(sale.final_amount for sale in weekly_sales) if weekly_sales else 0,
        'low_stock_count': Product.objects.filter(quantity__lte=models.F('min_quantity'), is_active=True).count(),
        'on_duty_today': WorkShift.objects.filter(date=today, is_active=True).count(),
        'categories': Category.objects.all().order_by('name'),
        'employees': Employee.objects.filter(is_active=True).order_by('last_name'),
    }
    return render(request, 'reports/dashboard.html', context)


def _generate_excel_report(data, report):
    """Генерация Excel отчета"""
    # Создаем новую книгу Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Стили
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Заголовок отчета
    ws.merge_cells('A1:D1')
    title_cell = ws['A1']
    title_cell.value = f"{report.get_report_type_display()} за период {report.start_date} - {report.end_date}"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center')
    
    # Основная информация
    ws['A3'] = 'Дата генерации:'
    ws['B3'] = timezone.now().strftime('%d.%m.%Y %H:%M')
    ws['A4'] = 'Период:'
    ws['B4'] = f"{report.start_date} - {report.end_date}"
    
    current_row = 6
    
    if data['report_type'] == 'financial':
        # Финансовый отчет
        ws['A6'] = 'Финансовая статистика:'
        ws['A6'].font = Font(bold=True)
        
        ws['A7'] = 'Всего продаж:'
        ws['B7'] = data['total_sales']
        ws['A8'] = 'Общая выручка:'
        ws['B8'] = data['total_amount']
        ws['B8'].number_format = '#,##0.00 ₽'
        ws['A9'] = 'Возвратов:'
        ws['B9'] = data['total_returns']
        ws['A10'] = 'Сумма возвратов:'
        ws['B10'] = data['return_amount']
        ws['B10'].number_format = '#,##0.00 ₽'
        ws['A11'] = 'Чистая выручка:'
        ws['B11'] = data['net_amount']
        ws['B11'].number_format = '#,##0.00 ₽'
        ws['A12'] = 'Прибыль:'
        ws['B12'] = data['total_profit']
        ws['B12'].number_format = '#,##0.00 ₽'
        ws['A13'] = 'Средний чек:'
        ws['B13'] = data['average_check']
        ws['B13'].number_format = '#,##0.00 ₽'
        
    elif data['report_type'] == 'products':
        # Отчет по товарам
        ws['A6'] = f'Отчет по товарам ({data["category"]}):'
        ws['A6'].font = Font(bold=True)
        
        ws['A7'] = 'Всего товаров:'
        ws['B7'] = data['total_products']
        ws['A8'] = 'Продано единиц:'
        ws['B8'] = data['total_products_sold']
        ws['A9'] = 'Общая выручка:'
        ws['B9'] = data['total_revenue']
        ws['B9'].number_format = '#,##0.00 ₽'
        ws['A10'] = 'Общая прибыль:'
        ws['B10'] = data['total_profit']
        ws['B10'].number_format = '#,##0.00 ₽'
        
        current_row = 12
        
        if data.get('products'):
            ws[f'A{current_row}'] = 'Товары:'
            ws[f'A{current_row}'].font = Font(bold=True)
            current_row += 1
            
            headers = ['Артикул', 'Наименование', 'Категория', 'Продано', 'Выручка', 'Прибыль', 'Остаток']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            
            current_row += 1
            
            for product in data['products']:
                ws.cell(row=current_row, column=1, value=product['sku'])
                ws.cell(row=current_row, column=2, value=product['name'])
                ws.cell(row=current_row, column=3, value=product['category'])
                ws.cell(row=current_row, column=4, value=product['quantity_sold'])
                ws.cell(row=current_row, column=5, value=product['revenue']).number_format = '#,##0.00 ₽'
                ws.cell(row=current_row, column=6, value=product['profit']).number_format = '#,##0.00 ₽'
                ws.cell(row=current_row, column=7, value=product['current_stock'])
                
                if product['is_low_stock']:
                    for col in range(1, 8):
                        ws.cell(row=current_row, column=col).fill = PatternFill(
                            start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
                        )
                
                current_row += 1
    
    elif data['report_type'] == 'work_schedule':
        # График работы
        ws['A6'] = f'График работы ({data["employee"]}):'
        ws['A6'].font = Font(bold=True)
        
        ws['A7'] = 'Всего сотрудников:'
        ws['B7'] = data['total_employees']
        ws['A8'] = 'Всего смен:'
        ws['B8'] = data['total_shifts']
        ws['A9'] = 'Всего часов:'
        ws['B9'] = data['total_hours']
        ws['B9'].number_format = '0.0'
        
        current_row = 11
        
        if data.get('employees'):
            ws[f'A{current_row}'] = 'Сотрудники:'
            ws[f'A{current_row}'].font = Font(bold=True)
            current_row += 1
            
            headers = ['ФИО', 'Должность', 'Кол-во смен', 'Часов']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            
            current_row += 1
            
            for employee in data['employees']:
                ws.cell(row=current_row, column=1, value=employee['full_name'])
                ws.cell(row=current_row, column=2, value=employee['position'])
                ws.cell(row=current_row, column=3, value=employee['shifts_count'])
                ws.cell(row=current_row, column=4, value=employee['total_hours']).number_format = '0.0'
                current_row += 1
    
    # Настраиваем ширину колонок
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Создаем HTTP-ответ с файлом Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"{report.get_report_type_display()}_{report.start_date}_{report.end_date}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
def quick_financial_report(request):
    """Быстрый финансовый отчет"""
    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        
        # Преобразуем строки в даты
        from datetime import datetime
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        # Создаем временный объект отчета
        report = Report(
            report_type='financial',
            start_date=start_date_obj,
            end_date=end_date_obj,
            generated_by=request.user
        )
        
        data = report.generate_data()
        return _generate_excel_report(data, report)
    
    return redirect('reports:reports_dashboard')


@login_required
def quick_products_report(request):
    """Быстрый отчет по товарам"""
    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        category_id = request.POST.get('category')
        
        # Преобразуем строки в даты
        from datetime import datetime
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        report = Report(
            report_type='products',
            start_date=start_date_obj,
            end_date=end_date_obj,
            generated_by=request.user
        )
        
        if category_id:
            from products.models import Category
            try:
                report.category = Category.objects.get(id=category_id)
            except Category.DoesNotExist:
                pass
        
        data = report.generate_data()
        return _generate_excel_report(data, report)
    
    return redirect('reports:reports_dashboard')


@login_required
def quick_schedule_report(request):
    """Быстрый отчет по графику"""
    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        employee_id = request.POST.get('employee')
        
        # Преобразуем строки в даты
        from datetime import datetime
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        report = Report(
            report_type='work_schedule',
            start_date=start_date_obj,
            end_date=end_date_obj,
            generated_by=request.user
        )
        
        if employee_id:
            from staff.models import Employee
            try:
                report.employee = Employee.objects.get(id=employee_id)
            except Employee.DoesNotExist:
                pass
        
        data = report.generate_data()
        return _generate_excel_report(data, report)
    
    return redirect('reports:reports_dashboard')


@login_required
def quick_daily_report(request):
    """Быстрый отчет за день"""
    if request.method == 'POST':
        date_str = request.POST.get('date')
        
        # Преобразуем строки в даты
        from datetime import datetime
        date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
        
        report = Report(
            report_type='daily',
            start_date=date_obj,
            end_date=date_obj,
            generated_by=request.user
        )
        
        data = report.generate_data()
        return _generate_excel_report(data, report)
    
    return redirect('reports:reports_dashboard')